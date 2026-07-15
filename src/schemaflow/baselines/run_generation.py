import os
import sys
import re
import json
import random
import logging
from tqdm import tqdm
from datetime import datetime
from openpyxl import Workbook, load_workbook

import gc
import torch
from json_repair import repair_json
from dataclasses import dataclass, field
from transformers import (
    HfArgumentParser,
    AutoTokenizer,
    BitsAndBytesConfig,
    Qwen3ForCausalLM,
    Qwen3_5ForCausalLM,
    GenerationConfig,
)
from accelerate.utils import set_seed
from accelerate import PartialState

from schemaflow.utils.io import setup_logger
from schemaflow.data.schemas import SCHEMAS

MODEL_CLASSES = {
    "qwen3": (Qwen3ForCausalLM, AutoTokenizer),
    "qwen3_5": (Qwen3_5ForCausalLM, AutoTokenizer),
}

ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
logger = None


def parse_llm_output(text: str):
    think_match = re.search(r"<think>(.*?)</think>", text, re.DOTALL)
    think_content = think_match.group(1).strip() if think_match else ""
    final_output = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()
    return think_content, final_output


def extract_json_str(text: str) -> str:
    match = re.search(r"\{.*\}", text.strip(), re.MULTILINE | re.IGNORECASE | re.DOTALL)
    if not match:
        raise ValueError(f"Could not extract json string from output: {text}")
    return match.group()


@dataclass
class ModelArguments:
    """
    Configuration class for the models.

    Using [`~transformers.HfArgumentParser`] we can turn this class into
    [argparse](https://docs.python.org/3/library/argparse#module-argparse) arguments that can be specified on the
    command line.

    bnb_4bit_compute_dtype (`torch.dtype` or str, *optional*, defaults to `torch.float32`):
            This sets the computational type which might be different than the input type. For example, inputs might be
            fp32, but computation can be set to bf16 for speedups.
    bnb_4bit_quant_type (`str`,  *optional*, defaults to `"fp4"`):
        This sets the quantization data type in the bnb.nn.Linear4Bit layers. Options are FP4 and NF4 data types
        which are specified by `fp4` or `nf4`.
    bnb_4bit_use_double_quant (`bool`, *optional*, defaults to `False`):
        This flag is used for nested quantization where the quantization constants from the first quantization are
        quantized again.

    """

    model_type: str = field(
        default=None,
        metadata={
            "help": "Model type selected in the list: "
            + ", ".join(MODEL_CLASSES.keys())
        },
    )

    model_name_or_path: str | None = field(
        default=None,
        metadata={"help": "Model checkpoint for weights initialization."},
    )
    model_revision: str = field(
        default="main",
        metadata={
            "help": "Specific model version to use. It can be a branch name, a tag name, or a commit id."
        },
    )
    dtype: str | None = field(
        default="auto",
        metadata={
            "help": "Override the default `torch.dtype` and load the model under this dtype. It defaults to `'float32'`.",
            "choices": ["auto", "bfloat16", "float16", "float32"],
        },
    )
    attn_implementation: str | None = field(
        default=None,
        metadata={
            "help": "Which attention implementation to use. You can run `--attn_implementation=flash_attention_2`, in "
            "which case you must install this manually by running `pip install flash-attn --no-build-isolation`."
        },
    )
    use_cpu: bool = field(
        default=False,
        metadata={
            "help": "Whether or not to use cpu. If set to False, we will use gpu/npu or mps device if available"
        },
    )
    load_in_8bit: bool = field(
        default=False,
        metadata={
            "help": "Whether to use 8 bit precision for the base model. Works only with LoRA."
        },
    )
    load_in_4bit: bool = field(
        default=False,
        metadata={
            "help": "Whether to use 4 bit precision for the base model. Works only with LoRA."
        },
    )
    bnb_4bit_quant_type: str = field(
        default="nf4",
        metadata={"help": "Quantization type.", "choices": ["fp4", "nf4"]},
    )
    use_bnb_nested_quant: bool = field(
        default=False,
        metadata={"help": "Whether to use nested quantization."},
    )
    bnb_4bit_compute_dtype: str | None = field(
        default=None,
        metadata={"help": "Quantization computational type"},
    )

    bnb_4bit_use_double_quant: bool = field(
        default=False, metadata={"help": "Flag used for nested quantization"}
    )

    def __post_init__(self):
        if self.load_in_8bit and self.load_in_4bit:
            raise ValueError("You can't use 8 bit and 4 bit precision at the same time")


@dataclass
class DataArguments:
    """
    Configuration class for the models.

    Using [`~transformers.HfArgumentParser`] we can turn this class into
    [argparse](https://docs.python.org/3/library/argparse#module-argparse) arguments that can be specified on the
    command line.

    """

    input_filepath: str = field(
        default=None,
        metadata={"help": "Path to the input filepath"},
    )
    output_filepath: str = field(
        default="outputs/output.xlsx",
        metadata={"help": "Path to the output filepath"},
    )
    logs_dir: str = field(
        default="logs/",
        metadata={"help": "Path to the logs directory"},
    )

    def __post_init__(self):
        if self.input_filepath is None:
            raise ValueError("input_filepath must be provided.")

        output_dir = os.path.dirname(self.output_filepath)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        os.makedirs(self.logs_dir, exist_ok=True)


@dataclass
class InferenceArguments:
    batch_size: int = field(
        default=8, metadata={"help": "Batch size for batched inference"}
    )
    max_new_tokens: int = field(
        default=10000, metadata={"help": "Maximum new tokens for text generation"}
    )
    do_sample: bool = field(default=True, metadata={"help": "Whether to use sampling"})
    temperature: float = field(default=1.0, metadata={"help": "Sampling temperature"})
    top_p: float = field(default=0.95, metadata={"help": "Top-p sampling"})
    top_k: int = field(default=20, metadata={"help": "Top-k sampling"})
    min_p: float = field(default=0.0, metadata={"help": "Min-p sampling"})
    presence_penalty: float = field(default=1.5, metadata={"help": "Presence penalty"})
    repetition_penalty: float = field(
        default=1.0, metadata={"help": "Repetition penalty"}
    )
    seed: int = field(default=42, metadata={"help": "random seed for initialization"})
    prompt_template: str = field(
        default="Schema:\n{schema}\n\nQuery:\n{query}\n\nReturn only valid JSON.",
        metadata={"help": "Prompt template"},
    )


def load_model(model_args: ModelArguments):
    if model_args.load_in_4bit:
        quantization_config = BitsAndBytesConfig(
            load_in_4bit=model_args.load_in_4bit,
            bnb_4bit_compute_dtype=model_args.bnb_4bit_compute_dtype,
            bnb_4bit_use_double_quant=model_args.bnb_4bit_use_double_quant,
            bnb_4bit_quant_type=model_args.bnb_4bit_quant_type,
        )
    elif model_args.load_in_8bit:
        quantization_config = BitsAndBytesConfig(load_in_8bit=True)
    else:
        quantization_config = None

    try:
        model_args.model_type = model_args.model_type.lower()
        model_class, tokenizer_class = MODEL_CLASSES[model_args.model_type]
    except KeyError:
        raise KeyError(
            f"the model {model_args.model_type} you specified is not supported."
        )

    tokenizer = tokenizer_class.from_pretrained(model_args.model_name_or_path)
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token

    model = model_class.from_pretrained(
        model_args.model_name_or_path, quantization_config=quantization_config
    )

    return model, tokenizer


def generate(
    messages: list[str],
    model,
    tokenizer,
    inference_args,
) -> list[str]:
    tokenizer.padding_side = "left"

    batch_size = inference_args.batch_size

    i = 0
    all_results = []
    while i < len(messages):
        batch = messages[i : i + batch_size]
        conversations = [[{"role": "user", "content": message}] for message in batch]

        try:
            inputs = tokenizer.apply_chat_template(
                conversations,
                add_generation_prompt=True,
                tokenize=True,
                return_dict=True,
                return_tensors="pt",
                padding=True,
                truncation=True,
            ).to(model.device)

            generation_config = GenerationConfig(
                max_new_tokens=inference_args.max_new_tokens,
                do_sample=inference_args.do_sample,
                temperature=inference_args.temperature,
                top_p=inference_args.top_p,
                top_k=inference_args.top_k,
                min_p=inference_args.min_p,
                repetition_penalty=inference_args.repetition_penalty,
                presence_penalty=inference_args.presence_penalty,
                pad_token_id=tokenizer.eos_token_id,
            )

            out_tokens = model.generate(
                **inputs,
                generation_config=generation_config,
            )
            
            prompt_length = inputs["input_ids"].shape[1]
            decoded = [
                tokenizer.decode(output[prompt_length:], skip_special_tokens=True,)
                for output in out_tokens
            ]

            all_results.extend(decoded)

            i += len(decoded)

        except (torch.cuda.OutOfMemoryError, RuntimeError) as e:
            torch.cuda.empty_cache()
            gc.collect()

            if torch.cuda.is_available():
                torch.cuda.reset_peak_memory_stats()

            logger.error(f"CUDA error: {e}")

            if batch_size == 1:
                logger.error("Skipping prompt")
                all_results.append(f"CUDA error: {e}")
                i += 1
            else:
                batch_size = max(1, batch_size // 2)
                logger.warning(f"Reducing batch size to {batch_size} and retrying.")

    return all_results


def main():
    parser = HfArgumentParser((ModelArguments, DataArguments, InferenceArguments))

    if len(sys.argv) == 1:
        parser.print_help()
        print(
            "\nProvide either:\n"
            "  1. A JSON config file:\n"
            "       python run_generation.py config.json\n\n"
            "  2. Command-line arguments:\n"
            "       python run_generation.py "
            "--model_type qwen3 "
            "--model_name_or_path Qwen/Qwen3-8B "
            "--filepath data.json"
        )
        return

    # Support a JSON config file
    if len(sys.argv) == 2 and sys.argv[1].endswith(".json"):
        model_args, data_args, inference_args = parser.parse_json_file(
            json_file=sys.argv[1]
        )
    else:
        model_args, data_args, inference_args = parser.parse_args_into_dataclasses()

    if inference_args.seed is not None:
        set_seed(inference_args.seed)

    logs_dir = data_args.logs_dir
    log_filepath = os.path.join(logs_dir, f"{ts_str}.log")

    global logger
    logger = setup_logger(log_filepath=log_filepath)

    distributed_state = PartialState(cpu=model_args.use_cpu)

    print(f"DEVICE: {distributed_state.device}")

    model, tokenizer = load_model(model_args=model_args)
    model.to(distributed_state.device)

    with open(data_args.input_filepath, "r", encoding="utf-8") as f:
        ds = json.load(f)

    prompt_template = inference_args.prompt_template

    messages = []
    for example in ds:
        schema = SCHEMAS[example["schema"]].describe()
        query = example["query"]
        messages.append(prompt_template.format(schema=schema, query=query))
    
    batch_size = inference_args.batch_size
    
    
    columns = [
        "query",
        "schema",
        "gold_cypher",
        "gold",
        "raw_llm_output",
        "llm_think",
        "llm_output",
        "json_parsed",
    ]
    
    output_path = data_args.output_filepath
    
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(columns)
        
    rows = []
    for start in tqdm(range(0, len(messages), batch_size), desc="Training",):
        end = start + batch_size

        batch_inputs = messages[start:end]
        batch_examples = ds[start:end]

        batch_results = generate(
            messages=batch_inputs,
            model=model,
            tokenizer=tokenizer,
            inference_args=inference_args,
        )
        
        for sample, model_output in zip(batch_examples, batch_results):
            think_content, response = parse_llm_output(model_output)

            json_parsed = None
            try:
                json_str = extract_json_str(response)
                json_parsed = repair_json(
                    json_str,
                    ensure_ascii=False,
                    return_objects=True,
                )
            except (ValueError, KeyError, TypeError) as e:
                logger.error(f"Error parsing JSON: {e}. Skipping JSON parsing.")
            
            logger.info(f"Query: {sample['query']}")
            logger.info(f"LLM Response: {json_parsed if json_parsed is not None else 'Invalid LLM Response'}")
            
            rows.append([
                sample["query"],
                sample["schema"],
                sample.get("gold_cypher"),
                json.dumps(sample.get("gold"), ensure_ascii=False),
                model_output,
                think_content,
                response,
                (
                    json.dumps(json_parsed, ensure_ascii=False)
                    if json_parsed is not None
                    else None
                ),
            ])

        for row in rows:
            ws.append(row)

        wb.save(output_path)

        logger.info(
            f"Processed batch {start // batch_size + 1} "
            f"({min(end, len(ds))}/{len(ds)} examples). "
            f"Saved {len(rows)} rows."
        )
        rows.clear()
        
        del batch_results
        gc.collect()
        torch.cuda.empty_cache()
        
    wb.close()

if __name__ == "__main__":
    main()
